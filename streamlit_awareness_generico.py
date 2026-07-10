# -*- coding: utf-8 -*-
"""
Generador Flexible de Awareness / Recordación - Streamlit

Funciona para:
- Bancos
- Conglomerados Financieros
- Marcas / Empresas / Productos / Personalizado
"""

import io
import re
import json
import unicodedata
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIONES PREDEFINIDAS
# ============================================================

# --- PRESET 1: BANCOS ---
BANKS = [
    "Banco Agrario", "Bancolombia", "Davivienda", "Banco de Bogotá",
    "BBVA", "Banco Caja Social", "Banco Popular", "Bancamía",
    "Banco de Occidente", "Banco Mundo Mujer",
]

BANK_ALIASES = {
    "Banco Agrario": ["banco agrario", "banco agrario de colombia", "agrario"],
    "Bancolombia": ["bancolombia"],
    "Davivienda": ["davivienda"],
    "Banco de Bogotá": ["banco de bogota", "banco de bogotá", "banco bogota", "banco bogotá"],
    "BBVA": ["bbva", "bbvva", "bvva", "bbwa", "bva", "bvvwa"],
    "Banco Caja Social": ["banco caja social", "caja social", "cajas social", "caja sosial"],
    "Banco Popular": ["banco popular", "popular"],
    "Bancamía": ["bancamia", "bancamía", "banca mia", "banca mía"],
    "Banco de Occidente": ["banco de occidente", "occidente"],
    "Banco Mundo Mujer": ["banco mundo mujer", "mundo mujer", "banco de la mujer", "fundacion de la mujer", "de la mujer"],
}

BANK_NORMALIZATIONS = [
    {"pattern": r"(?i).*Banco\s+de\s+Bogota.*", "replacement": "Banco de Bogotá"},
    {"pattern": r"(?i).*(?<!Banco\s)Caja\s+Social.*", "replacement": "Banco Caja Social"},
    {"pattern": r"(?i).*Bancamia.*", "replacement": "Bancamía"},
]

# --- PRESET 2: CONGLOMERADOS FINANCIEROS Y MARCAS GENERALES ---
CONGLOMERATES = [
    "Conglomerado BBVA", "Grupo cooperativo Coomeva", "Fundación Grupo Social",
    "Grupo Bolivar", "Conglomerado financiero Sura-Bancolombia", "Grupo Aval",
    "GNB Sudameris", "Conglomerado Credicorp capital"
]

CONGLOMERATE_ALIASES = {
    "Conglomerado BBVA": ["conglomerado bbva", "bbva"],
    "Grupo cooperativo Coomeva": ["grupo cooperativo coomeva", "coomeva"],
    "Fundación Grupo Social": ["fundacion grupo social", "grupo social"],
    "Grupo Bolivar": ["grupo bolivar", "bolivar"],
    "Conglomerado financiero Sura-Bancolombia": ["conglomerado financiero sura bancolombia", "sura", "bancolombia"],
    "Grupo Aval": ["grupo aval", "aval"],
    "GNB Sudameris": ["gnb sudameris", "sudameris", "gnb"],
    "Conglomerado Credicorp capital": ["conglomerado credicorp capital", "credicorp"]
}

CONGLOMERATE_NORMALIZATIONS = [
    {"pattern": r"(?i).*grupo\s+bancolombia.*", "replacement": "Grupo Bancolombia"},
    {"pattern": r"(?i).*(?<!grupo\s)bancolombia.*", "replacement": "Bancolombia"},
    {"pattern": r"(?i).*(exito|éxito).*", "replacement": "Grupo exito"},
    {"pattern": r"(?i).*ecopetrol.*", "replacement": "Grupo Ecopetrol"},
    {"pattern": r"(?i).*nutresa.*", "replacement": "Grupo nutresa"},
    {"pattern": r"(?i).*(sudameris|gnb).*", "replacement": "GNB Sudameris"},
    {"pattern": r"(?i).*credicorp.*", "replacement": "Conglomerado Credicorp capital"},
    {"pattern": r"(?i).*skandia.*", "replacement": "Conglomerado Skandia"},
    {"pattern": r"(?i).*(adidas|afidas|asidas|adida).*", "replacement": "Adidas"},
    {"pattern": r"(?i).*(coca-cola|cocacola|coca\s+cola).*", "replacement": "Coca-Cola"},
    {"pattern": r"(?i).*(av\s+villas|avvillas).*", "replacement": "Banco AV Villas"},
    {"pattern": r"(?i).*(nestle|nestlé).*", "replacement": "Nestlé"},
    {"pattern": r"(?i).*compensar.*", "replacement": "Compensar"}
]

NEGATIVOS = {
    "", "0", "no", "nan", "none", "false", "ninguno", "ninguna",
    "no se", "no sé", "no recuerdo", "ningun otro", "ningún otro", "no aplica",
}


# ============================================================
# UTILIDADES DE TEXTO
# ============================================================

def norm(value) -> str:
    if pd.isna(value): return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def clean_col(value) -> str:
    if pd.isna(value): return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip()

def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", str(name)).strip()
    return cleaned[:31] or "Hoja"

def apply_normalizations(value, normalizations: List[Dict]) -> str:
    if pd.isna(value):
        return value
    text = str(value).strip()
    for item in normalizations or []:
        pattern = item.get("pattern", "")
        replacement = item.get("replacement", "")
        if not pattern: continue
        try:
            if re.search(pattern, text):
                return replacement
        except re.error:
            pass
    return text

def contains_entity(value, entity: str, aliases: Dict[str, List[str]]) -> bool:
    text = norm(value)
    if text in NEGATIVOS: return False
    for alias in aliases.get(entity, [entity]):
        alias_norm = norm(alias)
        if alias_norm and re.search(r"(^|\s)" + re.escape(alias_norm) + r"(\s|$)", text):
            return True
    return False


# ============================================================
# DETECCIÓN DE COLUMNAS
# ============================================================

def find_col(df: pd.DataFrame, prefix: str) -> Optional[str]:
    prefix_norm = clean_col(prefix)
    if not prefix_norm or prefix_norm == "0": return None
    for column in df.columns:
        col_norm = clean_col(column)
        if prefix_norm in col_norm: 
            return column
    return None

def prefixes_to_list(text: str) -> List[str]:
    lines = str(text).replace(",", "\n").splitlines()
    return [line.strip() for line in lines if line.strip()]

def find_cols(df: pd.DataFrame, prefixes_text: str) -> List[str]:
    detected = []
    prefixes = [clean_col(p) for p in prefixes_to_list(prefixes_text) if p.strip() != "0"]
    if not prefixes: return detected
    for column in df.columns:
        col_norm = clean_col(column)
        if any(prefix in col_norm for prefix in prefixes):
            if column not in detected:
                detected.append(column)
    return detected

def entity_from_aided_col(column_name: str, entities: List[str]) -> Optional[str]:
    text = norm(str(column_name).replace("\n", " "))
    for entity in entities:
        if norm(entity) in text: return entity
    return None


# ============================================================
# CÁLCULO DE AWARENESS Y TABLAS
# ============================================================

def build_analysis(df: pd.DataFrame, demo_cols: Dict, cfg: Dict, entities: List[str], aliases: Dict, normalizations: List[Dict], expected_raw_cols: int):
    tom_col = find_col(df, cfg["tom"])
    esp_cols = find_cols(df, cfg["esp"])
    ayud_cols = find_cols(df, cfg["ayu"])

    raw_cols = [demo_cols.get(k) for k in ["sexo", "edad", "departamento", "estrato", "ingreso"]] + [tom_col] + esp_cols + ayud_cols
    raw_cols = [col for col in raw_cols if col is not None]

    if expected_raw_cols and expected_raw_cols > 0 and len(raw_cols) != expected_raw_cols:
        raise ValueError(f"Se esperaban {expected_raw_cols} columnas, pero se detectaron {len(raw_cols)}. Pon 0 en 'Columnas crudas esperadas'.")

    raw_df = df[raw_cols].copy()
    output_df = raw_df.copy()
    indicators = {}

    aided_map = {entity: None for entity in entities}
    for col in ayud_cols:
        entity = entity_from_aided_col(col, entities)
        if entity in aided_map: aided_map[entity] = col

    for entity in entities:
        tom = raw_df[tom_col].apply(lambda value: contains_entity(value, entity, aliases)) if tom_col else pd.Series(False, index=df.index)
        
        espontaneo = pd.Series(False, index=df.index)
        for col in esp_cols:
            espontaneo = espontaneo | raw_df[col].apply(lambda value: contains_entity(value, entity, aliases))

        aided_col = aided_map.get(entity)
        if aided_col is not None:
            def check_aided_value(value, ent, als):
                if pd.isna(value): return False
                v_str = str(value).strip().lower()
                if v_str in ["1", "si", "sí", "x", "seleccionado", "true"] or ent.lower() in v_str: return True
                return contains_entity(value, ent, als)
            ayudado = raw_df[aided_col].apply(lambda value: check_aided_value(value, entity, aliases))
        else:
            ayudado = pd.Series(False, index=df.index)
            for col in ayud_cols:
                ayudado = ayudado | raw_df[col].apply(lambda value: contains_entity(value, entity, aliases))

        indicators[(entity, "TOM")] = tom.astype(int)
        indicators[(entity, "Espontaneo")] = ((~tom) & espontaneo).astype(int)
        indicators[(entity, "Ayudado")] = ((~tom) & (~espontaneo) & ayudado).astype(int)

        output_df[(entity, "TOM")] = indicators[(entity, "TOM")]
        output_df[(entity, "Espontaneo")] = indicators[(entity, "Espontaneo")]
        output_df[(entity, "Ayudado")] = indicators[(entity, "Ayudado")]

    return output_df, pd.DataFrame(indicators), raw_cols


def tom_frequency_table(df: pd.DataFrame, tom_col: str, normalizations: List[Dict]) -> pd.DataFrame:
    if not tom_col:
        return pd.DataFrame()
    
    all_mentions = []
    
    for row_value in df[tom_col].dropna():
        val_str = str(row_value).strip()
        if val_str.lower() in NEGATIVOS or val_str == "":
            all_mentions.append("Ninguno / NS-NR")
            continue
        
        pieces = re.split(r'[,;]|\s+y\s+|\s+Y\s+', val_str)
        
        for piece in pieces:
            piece = piece.strip()
            if not piece or piece.lower() in NEGATIVOS:
                continue
            
            normalized_piece = apply_normalizations(piece, normalizations)
            v = str(normalized_piece).strip()
            if not (v.startswith("Conglomerado") or v.startswith("Grupo") or v.startswith("Fundación") or v.startswith("Banco")):
                v = v.title()  
                
            all_mentions.append(v)
            
    if not all_mentions:
        return pd.DataFrame(columns=["Marca / Categoría TOM", "Cantidad", "%"])
        
    freq = pd.Series(all_mentions).value_counts().reset_index()
    freq.columns = ["Marca / Categoría TOM", "Cantidad"]
    total = freq["Cantidad"].sum()
    freq["%"] = freq["Cantidad"] / total if total > 0 else 0
    return freq


def category_table(indicators, df, category_col, category_value, entities):
    mask = pd.Series(True, index=df.index) if category_col is None else df[category_col].fillna("Sin dato") == category_value
    base = int(mask.sum())
    rows = []
    for entity in entities:
        tom = int(indicators[(entity, "TOM")][mask].sum()) if (entity, "TOM") in indicators else 0
        esp = int(indicators[(entity, "Espontaneo")][mask].sum()) if (entity, "Espontaneo") in indicators else 0
        ayu = int(indicators[(entity, "Ayudado")][mask].sum()) if (entity, "Ayudado") in indicators else 0
        awa = tom + esp + ayu
        rows.append({"Entidad": entity, "TOM": tom, "TOM %": tom/base if base else 0, "Espontaneo": esp, "Espontaneo %": esp/base if base else 0, "Ayudado": ayu, "Ayudado %": ayu/base if base else 0, "AWA": awa, "AWA %": awa/base if base else 0})
    return pd.DataFrame(rows)

def awareness_sections(indicators, df, demo_cols, analysis_name, entities):
    sections = [(f"Resumen general {analysis_name}", category_table(indicators, df, None, None, entities))]
    for key in ["sexo", "estrato"]:
        col = demo_cols.get(key)
        if col:
            for value in pd.Series(df[col].fillna("Sin dato")).drop_duplicates():
                sections.append((f"Detalle por {key} - {value}", category_table(indicators, df, col, value, entities)))
    return sections

def department_sections(indicators, df, demo_cols, entities):
    dept_col = demo_cols.get("departamento")
    if not dept_col: return [("Sin departamento/ciudad detectado", pd.DataFrame())]
    return [(str(value), category_table(indicators, df, dept_col, value, entities)) for value in pd.Series(df[dept_col].fillna("Sin dato")).drop_duplicates()]

def demographic_sections(df, demo_cols):
    sections = [("Base", pd.DataFrame({"Indicador": ["Total encuestas procesadas"], "Valor": [len(df)]}))]
    for key, label in [("sexo", "Sexo"), ("edad", "Edad"), ("departamento", "Ciudad/Depto"), ("estrato", "Estrato"), ("ingreso", "Ingreso")]:
        col = demo_cols.get(key)
        if col:
            table = df[col].fillna("Sin dato").value_counts().rename_axis(label).reset_index(name="Cantidad")
            table["%"] = table["Cantidad"] / len(df) if len(df) else 0
            sections.append((f"Demográfico - {label}", table))
    return sections

def style_sheet(ws):
    border = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"), top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

def write_main_sheet(wb, sheet_name, output_df):
    ws = wb.create_sheet(safe_sheet_name(sheet_name))
    for col_idx, column in enumerate(output_df.columns, start=1):
        ws.cell(1, col_idx).value = str(column)
        ws.cell(1, col_idx).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(1, col_idx).font = Font(bold=True)
    for row_idx, row in enumerate(output_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx).value = value
    style_sheet(ws)

def write_sections_sheet(wb, sheet_name, sections):
    ws = wb.create_sheet(safe_sheet_name(sheet_name))
    row_idx = 1
    for title, df_table in sections:
        ws.cell(row_idx, 1).value = title
        ws.cell(row_idx, 1).fill = PatternFill("solid", fgColor="C6E0B4")
        ws.cell(row_idx, 1).font = Font(bold=True)
        row_idx += 1
        if df_table.empty:
            ws.cell(row_idx, 1).value = "Sin datos disponibles"
            row_idx += 3
            continue
        for col_idx, column in enumerate(df_table.columns, start=1):
            ws.cell(row_idx, col_idx).value = column
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(row_idx, col_idx).font = Font(bold=True, color="FFFFFF")
        row_idx += 1
        for record in df_table.itertuples(index=False):
            for col_idx, value in enumerate(record, start=1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = value
                if str(df_table.columns[col_idx - 1]).endswith("%"):
                    cell.number_format = "0%"
            row_idx += 1
        row_idx += 2
    style_sheet(ws)

def build_excel_bytes(df, demo_cols, cfg1, cfg2, entities, aliases, normalizations, expected_raw_cols):
    run_1 = any(cfg1[k].strip() not in ["0", ""] for k in ["tom", "esp", "ayu"])
    run_2 = any(cfg2[k].strip() not in ["0", ""] for k in ["tom", "esp", "ayu"])
    
    if not run_1 and not run_2: 
        raise ValueError("Debe configurar al menos un análisis válido (TOM, Espontáneo o Ayudado). Escriba los prefijos de columnas en la sección 2.")
    
    wb = Workbook()
    wb.remove(wb.active)
    write_sections_sheet(wb, "Resumen Demográficos", demographic_sections(df, demo_cols))

    for cfg, run_flag in [(cfg1, run_1), (cfg2, run_2)]:
        if run_flag:
            output_df, indicators, _ = build_analysis(df, demo_cols, cfg, entities, aliases, normalizations, expected_raw_cols)
            write_main_sheet(wb, cfg["name"], output_df)
            write_sections_sheet(wb, f"Resumen {cfg['name']}", awareness_sections(indicators, df, demo_cols, cfg["name"], entities))
            write_sections_sheet(wb, f"Deptos {cfg['name']}", department_sections(indicators, df, demo_cols, entities))
            
            tom_col_name = find_col(df, cfg["tom"])
            if tom_col_name:
                freq_table = tom_frequency_table(df, tom_col_name, normalizations)
                write_sections_sheet(wb, f"Frecuencias P1 {cfg['name']}", [("Distribución de Respuestas P1 (TOM)", freq_table)])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ============================================================
# INTERFAZ STREAMLIT
# ============================================================

st.set_page_config(page_title="Awareness Flexible", layout="wide")
st.title("Generador flexible de Awareness / Recordación")

with st.sidebar:
    st.header("Configuración general")
    mode = st.radio("Tipo de estudio", ["Bancos", "Conglomerados Financieros", "Personalizado"], index=2)
    uploaded_file = st.file_uploader("Archivo base .xlsx o .csv", type=["xlsx", "csv"])
    max_rows = st.number_input("Filas a evaluar (0 = todas)", min_value=0, value=0, step=50)
    default_expected_cols = 17 if mode == "Bancos" else 0
    expected_raw_cols = st.number_input("Columnas crudas esperadas (0 = no validar)", min_value=0, value=default_expected_cols, step=1)

if mode == "Bancos":
    default_entities, default_aliases, default_normalizations = BANKS, BANK_ALIASES, BANK_NORMALIZATIONS
    def_sexo, def_edad, def_depto, def_estrato, def_ingreso = "F1 ", "F2 ", "F4 ", "F3 ", "0"
    t1_name, t1_tom, t1_esp, t1_ayu = "AWA PUB", "0", "0", "0"
    t2_name, t2_tom, t2_esp, t2_ayu = "AWA Marca", "P1", "P1A.1\nP1A.2\nP1A.3", "P2.1\nP2.2\nP2.3"
elif mode == "Conglomerados Financieros":
    default_entities, default_aliases, default_normalizations = CONGLOMERATES, CONGLOMERATE_ALIASES, CONGLOMERATE_NORMALIZATIONS
    def_sexo, def_edad, def_depto, def_estrato, def_ingreso = "F1", "F2a", "F4", "F3", "F5"
    t1_name, t1_tom, t1_esp, t1_ayu = "Desactivado", "0", "0", "0"
    t2_name, t2_tom, t2_esp, t2_ayu = "Conglomerados AWA", "P1 -", "P1A -", "P2-P2."
else:
    default_entities = ["Entidad 1"]
    default_aliases = {
        "Entidad 1": ["Entidad 1", "e1"],
       
    }
    default_normalizations = []
    def_sexo, def_edad, def_depto, def_estrato, def_ingreso = "", "", "", "", ""
    t1_name, t1_tom, t1_esp, t1_ayu = "Análisis 1", "0", "0", "0"
    t2_name, t2_tom, t2_esp, t2_ayu = "Análisis 2", "0", "0", "0"

with st.expander("1. Entidades, alias y normalizaciones", expanded=(mode == "Personalizado")):
    entities_text = st.text_area("Entidades a evaluar, una por línea", "\n".join(default_entities), height=200)
    aliases_text = st.text_area("Alias / condiciones en JSON", json.dumps(default_aliases, ensure_ascii=False, indent=2), height=220)
    normalizations_text = st.text_area("Normalizaciones opcionales en JSON", json.dumps(default_normalizations, ensure_ascii=False, indent=2), height=250)

try:
    aliases, normalizations = json.loads(aliases_text), json.loads(normalizations_text)
except Exception as exc:
    st.error(f"Error en estructuras JSON: {exc}"); st.stop()

entities = [line.strip() for line in entities_text.splitlines() if line.strip()]

with st.expander("2. Preguntas y demográficos", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        sexo_prefix = st.text_input("Sexo", def_sexo)
        edad_prefix = st.text_input("Edad", def_edad)
        departamento_prefix = st.text_input("Departamento / Ciudad", def_depto)
        estrato_prefix = st.text_input("Estrato", def_estrato)
        ingreso_prefix = st.text_input("Ingreso / Rangos", def_ingreso)
    with col2:
        a1_name = st.text_input("Nombre análisis 1", t1_name)
        a1_tom = st.text_input("TOM análisis 1", t1_tom)
        a1_esp = st.text_area("Espontáneo análisis 1", t1_esp, height=100)
        a1_ayu = st.text_area("Ayudado análisis 1", t1_ayu, height=100)
    with col3:
        a2_name = st.text_input("Nombre análisis 2", t2_name)
        a2_tom = st.text_input("TOM análisis 2", t2_tom)
        a2_esp = st.text_area("Espontáneo análisis 2", t2_esp, height=100)
        a2_ayu = st.text_area("Ayudado análisis 2", t2_ayu, height=100)

if uploaded_file is None:
    st.info("Por favor carga el archivo de datos para iniciar.")
    st.stop()

try:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file, engine="openpyxl")
except Exception as exc:
    st.error(f"Error al leer el archivo: {exc}"); st.stop()

if max_rows and max_rows > 0: df = df.head(int(max_rows)).copy()

demo_cols = {
    "sexo": find_col(df, sexo_prefix), "edad": find_col(df, edad_prefix),
    "departamento": find_col(df, departamento_prefix), "estrato": find_col(df, estrato_prefix),
    "ingreso": find_col(df, ingreso_prefix),
}

cfg1, cfg2 = {"name": a1_name, "tom": a1_tom, "esp": a1_esp, "ayu": a1_ayu}, {"name": a2_name, "tom": a2_tom, "esp": a2_esp, "ayu": a2_ayu}

st.subheader("3. Validación de Columnas")
validation_rows = [{"Tipo": "Demográfico", "Campo": k, "Columna detectada": str(v)} for k, v in demo_cols.items()]

for label, cfg in [("Análisis 1", cfg1), ("Análisis 2", cfg2)]:
    if any(cfg[k].strip() not in ["0", ""] for k in ["tom", "esp", "ayu"]):
        validation_rows.append({"Tipo": label, "Campo": "TOM", "Columna detectada": str(find_col(df, cfg["tom"]))})
        validation_rows.append({"Tipo": label, "Campo": "Espontáneo", "Columna detectada": ", ".join(find_cols(df, cfg["esp"]))})
        validation_rows.append({"Tipo": label, "Campo": "Ayudado", "Columna detectada": ", ".join(find_cols(df, cfg["ayu"]))})

st.dataframe(pd.DataFrame(validation_rows), use_container_width=True)

st.subheader("4. Procesar y Descargar")
if st.button("Generar Reporte Excel", type="primary"):
    try:
        result = build_excel_bytes(df, demo_cols, cfg1, cfg2, entities, aliases, normalizations, int(expected_raw_cols))
        st.success("¡Cálculo e indicadores generados exitosamente!")
        st.download_button(label="📥 Descargar Reporte de Awareness", data=result, file_name="Reporte_Awareness_Estructural.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as exc:
        st.error(f"Error: {exc}")
        st.exception(exc)
